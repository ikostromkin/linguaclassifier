from openpyxl import Workbook
from models import *


def create_report(research_id):
    try:
        research = Researches.get(id=research_id)
        if research:
            wb = Workbook()
            wb.remove(wb.active)

            ws1 = wb.create_sheet('Сводная таблица')
            ws2 = wb.create_sheet('Матрица')

            fields = research.fields.split(' ')
            fields = list(filter(None, fields))
            count = len(fields)

            material_list = research.material.split(' ')

            d = {}
            for i, word in enumerate(material_list):
                d[word] = i

            elements_list = []  # Список элементов по всему исследованию
            number_of_groups_list = []  # Список числа групп по респондентам

            column_names = fields
            for i in range(5-count):
                column_names.append('')
            column_names.append('Название группы')
            column_names.append('Элементы группы')
            ws1.append(column_names)

            respondents = Respondents.select().where(
                (Respondents.owner == research_id) & (Respondents.filled_all == True)).dicts()

            for respondent in respondents:
                groups = Groups.select().where(Groups.owner == respondent['id']).dicts()
                number_of_groups_list.append(len(groups))
                for group in groups:
                    temp = [
                        respondent['field1'],
                        respondent['field2'],
                        respondent['field3'],
                        respondent['field4'],
                        respondent['field5'],
                        group['name'],
                        group['elements']
                    ]
                    elements_list.append(group['elements'])
                    ws1.append(temp)

            num_elements_list = []

            for elements in elements_list:
                temp = elements.split(' ')
                temp = list(filter(None, temp))
                num_temp = []
                for word in temp:
                    num_temp.append(d[word])
                num_temp.sort()
                num_elements_list.append(num_temp)

            mat = []

            for i in range(50):
                l = []
                for j in range(50):
                    l.append(0)
                mat.append(l)

            for line in num_elements_list:
                for i in range(len(line) - 1):
                    for j in range(i + 1, len(line)):
                        mat[line[i]][line[j]] += 1

            for line in mat:
                ws2.append(line)

            wb.save('temp.xlsx')

    except DoesNotExist:
        return False
